from tkinter import *
from tkinter import messagebox
from datetime import datetime, timezone, timedelta
from openpyxl import load_workbook
import os
from win32com import client

dif = timedelta(hours=-3)
fuso=timezone(dif)
agora = datetime.now()
agora = agora.astimezone(fuso)
agora = agora.strftime('%d/%m/%Y %H:%M')
registro=""

janela = Tk()

def geraPDF():
    folder = "C:\\TR COVID"
    file_type = 'xlsx'
    out_folder = folder + "\\Laudos_PDF"

    os.chdir(folder)

    if not os.path.exists(out_folder):
        print('Creating output folder...')
        os.makedirs(out_folder)
        print(out_folder, 'created.')
    else:
        print(out_folder, 'already exists.\n')

    for files in os.listdir("."):
        if files.endswith(".xlsx"):
            print(files)

    print('\n\n')

    word = client.DispatchEx("Excel.Application")
    for files in os.listdir("."):
        if files.endswith(".xlsx") or files.endswith('xls'):
            out_name = files.replace(file_type, r"pdf")
            in_file = os.path.abspath(folder + "\\" + files)
            out_file = os.path.abspath(out_folder + "\\" + out_name)
            doc = word.Workbooks.Open(in_file)
            print('Exporting', out_file)
            doc.SaveAs(out_file, FileFormat=57)
            doc.Close()
  
def bt_save():
    col=nome.get()
    col=col.upper()
    nasc=dn.get()
    igm=str(v1.get())
    igg=str(v2.get())
    lote=str(v3.get())
    if col=="":
        messagebox.showinfo('ERRO','Preencha o nome!')
    else:
        if igm=='1':
            igm='REAGENTE'
        if igm=='2':
            igm='NÃO REAGENTE'
        if igg=='1':
            igg='REAGENTE'
        if igg=='2':
            igg='NÃO REAGENTE'
        if lote=='1':
            lote='Lote: 202005041'
            validade='Validade: 27/11/2020'
        if lote=='2':
            lote='Lote: 202006006'
            validade='Validade: 03/12/2020'
        indice = lista.curselection()[0]
        if indice == 0:
            resp='Analista 1'
            registro='CRBM - 12345'
            cargo='BIOMÉDICO'
        if indice == 1:
            resp='Analista 2'
            registro='CRBM - 12345'
            cargo='BIOMÉDICO'
        if indice == 2:
            resp='Analista 3'
            registro='CRBio - 12345'
            cargo='BIÓLOGO'
        if indice == 3:
            resp='Analista 4'
            registro='CRM - 12345'
            cargo='MÉDICO PATOLOGISTA'
        #gera excel
        caminho = 'laudo.xlsx'
        arquivo_excel = load_workbook(caminho)
        Planilha1 = arquivo_excel.active
        Planilha1['C3'] = col
        Planilha1['C4'] = nasc
        Planilha1['C6'] = igm
        Planilha1['C7'] = igg
        Planilha1['A13'] = lote
        Planilha1['A14'] = validade
        Planilha1['D22'] = resp+'-'+registro
        Planilha1['A23'] = agora
        Planilha1['D23'] = cargo
        arquivo_excel.save(col+".xlsx")
        myfile="C:\\TR COVID\\Laudos_PDF\\laudo.pdf"
        if os.path.isfile(myfile):
            os.remove(myfile)
        geraPDF()
        os.remove(myfile)
        os.remove("C:\\TR COVID\\"+col+".xlsx")
        

        caminho2 = 'controle.xlsx'
        arquivo_controle = load_workbook(caminho2)
        Plan1 = arquivo_controle.active
        lote2=lote[6:15]
        validade2=validade[10:20]
        valorControle = [agora, col, nasc, igm, igg, lote2, validade2, resp]
        Plan1.append(valorControle)
        arquivo_controle.save("controle.xlsx")
        myfile2="C:\\TR COVID\\Laudos_PDF\\controle.pdf"
        if os.path.isfile(myfile2):
            os.remove(myfile2)

        messagebox.showinfo('AVISO','Laudo de '+col+' gerado com sucesso!')

        #limpando campos
        nome.delete(0,END)
        dn.delete(0,END)
        lista.selection_set(0)
        v1.set(2)
        v2.set(2)
        v3.set(1)
   
#interface gráfica
txt1 = Label(janela, text="Nome completo:*")
txt1["font"]=("Arial", "10", "bold")
txt1.pack(side=LEFT)
txt1.place(x=8, y=10)

nome = Entry(janela)
nome["width"] = 65
nome.pack(side=LEFT)
nome.place(x=10, y=30)

txt2 = Label(janela, text="DN (dd/mm/aaa):")
txt2["font"]=("Arial", "10", "bold")
txt2.pack(side=LEFT)
txt2.place(x=447, y=10)

dn = Entry(janela)
dn["width"] = 15
dn.pack(side=LEFT)
dn.place(x=450, y=30)

txt3 = Label(janela, text="IgM:")
txt3["font"]=("Arial", "10", "bold")
txt3.pack(side=LEFT)
txt3.place(x=10, y=60)

v1 = IntVar()
v2 = IntVar()
v1.set(2)
v2.set(2)

reag = Radiobutton(janela, text='REAGENTE', variable=v1, value=1)
reag.place(x=40, y=60)

nreag = Radiobutton(janela, text='NÃO REAGENTE', variable=v1, value=2)
nreag.place(x=130, y=60)

txt4 = Label(janela, text="Lote:")
txt4["font"]=("Arial", "10", "bold")
txt4.pack(side=LEFT)
txt4.place(x=10, y=130)

v3 = IntVar()
v3.set(1)

lote5041 = Radiobutton(janela, text='202005041', variable=v3, value=1)
lote5041.place(x=45, y=130)

lote6006 = Radiobutton(janela, text='202006006', variable=v3, value=2)
lote6006.place(x=131, y=130)

txt5 = Label(janela, text="IgG:")
txt5["font"]=("Arial", "10", "bold")
txt5.pack(side=LEFT)
txt5.place(x=10, y=85)

reag2 = Radiobutton(janela, text='REAGENTE', variable=v2, value=1)
reag2.place(x=40, y=85)

nreag2 = Radiobutton(janela, text='NÃO REAGENTE', variable=v2, value=2)
nreag2.place(x=130, y=85)

txt6 = Label(janela, text="Responsável:")
txt6["font"]=("Arial", "10", "bold")
txt6.pack(side=LEFT)
txt6.place(x=300, y=60)

txt7 = Label(janela, text="Desenvolvido por Roberto Molina")
txt7["font"]=("Arial", "7", "italic")
txt7.pack(side=LEFT)
txt7.place(x=5, y=180)

profs = ["Analista 1", "Analista 2", "Analista 3", "Analista 4"]
lista = Listbox(janela)
for item in profs:
    lista.insert(END,item)
lista.selection_set(0)
lista.place(x=300, y=80)
lista["height"]=4

bt1 = Button(janela, width=20, text="GERAR LAUDO", command=bt_save)
bt1.place(x=450, y=120)

#Janela
janela.title("HOSPITAL FEDERAL - TESTE RÁPIDO PARA COVID-19")
janela.geometry("650x200")
janela.mainloop()
